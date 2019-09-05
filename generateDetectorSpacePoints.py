from openpyxl import load_workbook
import os
import json
import math


def loadDictionaryRowsFromWorkbook(path, headersAddress, dataAddress):
    wb = load_workbook(path)
    ws = wb.active
    headers = [header.value for row in ws[headersAddress] for header in row]

    nestedData = [list(zip(headers, [d.value for d in row]))
                  for row in ws[dataAddress]]

    data = []
    for row in nestedData:
        obj = {}
        for pair in row:
            obj[pair[0]] = pair[1]
        data.append(obj)

    return data


def createSupermoduleStackLayers(data, createPads=True):
    layers = [None for y in range(6)]

    for obj in data:
        layer = obj["layer"] = int(obj["type"][1])
        # Expect one C0 then corresponding C1
        if (obj["type"][3] == '0'):
            layers[layer] = obj.copy()
        else:
            # Create rows of type L?C1, L?C1, L?C0, L?C1, L?C1
            newLayer = [obj.copy() for x in range(5)]
            newLayer[2] = layers[obj["layer"]]
            # Enumerate stacks
            for stack in range(5):
                newLayer[stack]["stack"] = stack

            # Create pad planes
            createPadPlanes(newLayer, createPads)

            # Save completed layer
            layers[layer] = newLayer

    return layers


def createPadPlanes(modules, createPads=True):
    # Determine total length across all stacks
    lTotal = sum(map(lambda module: int(module["Lz"]), modules))
    l0 = int(-lTotal / 2)
    for module in modules:
        module["l0"] = l0
        module["l1"] = l0

        l0 = l0 + int(module["Lrim"])

        r0 = float(module["Rmin"])
        r1 = float(module["Rmax"])

        pads = []

        rows = module["rows"]
        for row in range(rows):
            l1 = l0
            if (row == 0 or row + 1 == rows):
                l1 = l1 + int(module["Lopad"])
            else:
                l1 = l1 + int(module["Lipad"])

            w0 = -(int(module["Wr"]) / 2)

            module["w0"] = round(w0, 2)

            w0 = w0 + (int(module["Wrim"]))
            for col in range(144):
                w1 = w0
                if (col == 0 or col + 1 == 144):
                    w1 = w0 + float(module["Wopad"])
                else:
                    w1 = w0 + float(module["Wipad"])

                pads.append({
                    "row": row,
                    "col": col,
                    "l0": l0,
                    "l1": l1,
                    "w0": round(w0, 2),
                    "w1": round(w1, 2),
                    "r0": r0,
                    "r1": r1
                })
                w0 = w1

            w0 = w0 + (int(module["Wrim"]))
            module["w1"] = round(w0, 2)

            l0 = l1

        l0 = l0 + int(module["Lrim"])

        module["l1"] = l0

        if (createPads):
            module["pads"] = pads


def outputJsonToFile(data, path):
    print("Writing: " + path)
    outfile = open(path, "w")
    json.dump(data, outfile, indent=4)
    outfile.close()

def outputJsonAsFunctionToFile(data, outfile, functionName):
    print("function " + functionName + "() { const dims = ", file = outfile, end = '')
    json.dump(data, outfile, indent = 4)
    print(";\n\n\treturn dims;\n}\n", file = outfile)


def rotate(point, degangle, origin=[0, 0]):
    """
    Rotate a point counterclockwise by a given angle around a given origin.

    The angle should be given in degrees.
    """
    ox, oy = origin
    px, py = point

    angle = degangle / 180 * math.pi

    qx = ox + math.cos(angle) * (px - ox) - math.sin(angle) * (py - oy)
    qy = oy + math.sin(angle) * (px - ox) + math.cos(angle) * (py - oy)

    return {
        "x": round(qx, 2), 
        "y": round(qy, 2)
    }


def geomSectorXYPlane(supermodule):
    sectors = []
    for sectorNumber in range(18):
        rotation = 10 + (20 * sectorNumber)
        for layerArray in supermodule:
            for module in layerArray:
                if (module["stack"] != 2):
                    continue

                x0 = module["Rmin"]
                x1 = module["Rmax"]

                y0 = module["w0"] / 10
                y1 = module["w1"] / 10
                ym = (y0 + y1) / 2

                sector = {
                    "sec": sectorNumber,
                    "lyr": module["layer"],
                    # Projected coordinates in clockwise-order
                    "d": [
                        rotate([x0, y0], rotation),
                        rotate([x0, ym], rotation),
                        rotate([x1, ym], rotation),
                        rotate([x0, ym], rotation),
                        rotate([x0, y1], rotation),
                        rotate([x1, y1], rotation),
                        rotate([x1, y0], rotation)
                    ]
                }

                sectors.append(sector)

    return sectors

def geomSectorXYPlaneZoom(supermodule):
    modules = []
    pads = []
    rotation = 90
    for layerArray in supermodule:
        for module in layerArray:
            if (module["stack"] != 2):
                continue

            x0 = module["Rmin"]
            x1 = module["Rmax"]

            y0 = module["w0"] / 10
            y1 = module["w1"] / 10

            sector = {
                "lyr": module["layer"],
                # Projected coordinates in clockwise-order
                "d": [
                    rotate([x0, y0], rotation),
                    rotate([x0, y1], rotation),
                    rotate([x1, y1], rotation),
                    rotate([x1, y0], rotation)
                ]
            }

            modules.append(sector)

            for pad in module["pads"]:
                if (pad["row"] != 1): continue

                py0 = pad["w0"] / 10
                py1 = pad["w1"] / 10
                padgeom = {
                    "l": module["layer"],
                    "c": pad["col"],
                    "d": [
                        rotate([x0, py0], rotation),
                        rotate([x0, py1], rotation),
                        rotate([x1, py1], rotation),
                        rotate([x1, py0], rotation)
                    ]
                }

                pads.append(padgeom)
                

    return modules, pads

def rid(stk, lyr, row):
   return (stk * 6 + lyr) * 16 + row

def geomStackZRPlane(supermodulePads):
    modules = []
    pads = []
    rotation = 0

    for layerArray in supermodulePads:
        for module in layerArray:
            y0 = module["Rmin"]
            y1 = module["Rmax"]

            x0 = module["l0"] / 10
            x1 = module["l1"] / 10

            moduleGeom = {
                "stk": module["stack"],
                "lyr": module["layer"],
                # Projected coordinates in clockwise-order
                "d": [
                    rotate([x0, y0], rotation),
                    rotate([x0, y1], rotation),
                    rotate([x1, y1], rotation),
                    rotate([x1, y0], rotation)
                ]
            }

            modules.append(moduleGeom)

            for pad in module["pads"]:
                if (pad["col"] != 1): continue

                px0 = pad["l0"] / 10
                px1 = pad["l1"] / 10
                padgeom = {
                    "rid": rid(module["stack"], module["layer"], pad["row"]),
                    "s": module["stack"],
                    "l": module["layer"],
                    "r": pad["row"],
                    "d": [
                        rotate([px0, y0], rotation),
                        rotate([px0, y1], rotation),
                        rotate([px1, y1], rotation),
                        rotate([px1, y0], rotation)
                    ]
                }

                pads.append(padgeom)

            # Create dummy entries for stacks with only 12 rather than 16 rows
            if (module["rows"] == 12):
                for i in range(4):
                    pads.append({
                        "rid": rid(module["stack"], module["layer"], 12 + i),
                        "s": module["stack"],
                        "l": module["layer"],
                        "r": pad["row"],
                        "d": []
                    })

    modules.sort(key = lambda x: x["stk"] * 10 + x["lyr"])
    pads.sort(key = lambda x: x["rid"])

    return modules, pads

def geomSectorXYPlaneTPC():
    outerPath = [rotate([0, 250], angle) for angle in range(0, 370, 10)]
    innerPath = [rotate([0, 85], angle) for angle in range(360, -10, -10)]

    return [outerPath, innerPath]

def geomSectorZRPlaneTPC():
    return [
        rotate([-250, 250], 0),
        rotate([250, 250], 0),
        rotate([250, 85], 0),
        rotate([-250, 85], 0),
        rotate([-250, 250], 0)
    ]

def geomLayers3D(supermodule):
    layers = []
    for sectorNumber in range(18):
        rotation = 10 + (20 * sectorNumber)
        for layerArray in supermodule:
            for module in layerArray:

                x0 = module["Rmin"]
                x1 = module["Rmax"]

                y0 = module["w0"] / 10
                y1 = module["w1"] / 10
                ym = (y0 + y1) / 2

                z0 = module["l0"] / 10
                z1 = module["l1"] / 10

                sector = {
                    "sec": sectorNumber,
                    "stk": module["stack"],
                    "lyr": module["layer"],
                    "rot": 20 * sectorNumber,

                    "w": abs(x1 - x0),
                    "h": abs(y1 - y0),
                    "d": abs(z1 - z0),

                    "x": (x0 + x1) / 2,
                    "y": (y0 + y1) / 2,
                    "z": (z0 + z1) / 2,
                }

                layers.append(sector)

    return layers

def generate():
    data = loadDictionaryRowsFromWorkbook(
        "jsroot/PadPlaneDimensions.xlsx", "A1:M1", "A2:M13")

    supermodule = createSupermoduleStackLayers(data, False)
    supermodulePads = createSupermoduleStackLayers(data, True)

    # outputJsonToFile(createSectorModules(supermodule),
    #                  "jsroot/geometry/sector-xy-plane.json")

    outfile = open("jsroot/geometry/geometries.js", "w")

    outputJsonAsFunctionToFile(geomSectorXYPlane(supermodule), outfile, "geomSectorXYPlane")

    modules, pads = geomSectorXYPlaneZoom(supermodulePads)
    outputJsonAsFunctionToFile(modules, outfile, "geomZoomSectorXYPlaneModules")
    outputJsonAsFunctionToFile(pads, outfile, "geomZoomSectorXYPlanePads")

    modules, pads = geomStackZRPlane(supermodulePads)
    outputJsonAsFunctionToFile(modules, outfile, "geomStackZRPlaneModules")
    outputJsonAsFunctionToFile(pads, outfile, "geomStackZRPlanePads")

    outputJsonAsFunctionToFile(geomSectorXYPlaneTPC(), outfile, "geomSectorXYPlaneTPC")
    outputJsonAsFunctionToFile(geomSectorZRPlaneTPC(), outfile, "geomSectorZRPlaneTPC")

    outfile.close()

    outfile = open("jsroot/geometry/geometries3d.js", "w")

    outputJsonAsFunctionToFile(geomLayers3D(supermodule), outfile, "geomLayers3D")

    print("export { geomLayers3D };", file = outfile)

    outfile.close()

    #outputJsonToFile(supermodule, "jsroot/geometry/supermodule.json")
    #outputJsonToFile(supermodulePads, "jsroot/geometry/supermodule-pads.json")

if __name__ == "__main__":
    generate()