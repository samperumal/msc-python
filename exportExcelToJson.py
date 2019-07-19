from openpyxl import load_workbook;
import os;
import json;
import math;

def main(headersAddress, dataAddress):
    wb = load_workbook("jsroot/TRDDimensions.xlsx")
    ws = wb.active
    headers = [header.value for row in ws[headersAddress] for header in row ]
    
    nestedData = [list(zip(headers, [d.value for d in row])) for row in ws[dataAddress]]
    data = []
    for row in nestedData:
        obj = {}
        for pair in row:
            obj[pair[0]] = pair[1]
        data.append(obj)

    outfile = open("jsroot/components/trd-dimensions.js", "w")
    print("function getDimensions() { const dims = ", file = outfile, end = '')
    json.dump(data, outfile, indent = 4)
    print("\t;\n\n\treturn dims;\n}", file = outfile)
    outfile.close()

    padrows = []
    for module in data:
        for row in range(16):
            used = row < module["zegments"]
            obj = {}
            obj["rid"] = (module["stack"] * 6 + module["layer"]) * 16 + row
            obj["use"] = used
            if (used):
                obj["stk"] = module["stack"]
                obj["lyr"] = module["layer"]
                obj["row"] = row
                obj["rows"] = module["zegments"]
                obj["p0"] = {
                    "r": module["minR"],
                    "z": round(module["minZ"] + row * module["zsize"], 2)
                }
                obj["p1"] = {
                    "r": module["maxR"],
                    "z": round(module["minZ"] + (row + 1) * module["zsize"], 2)
                }
            padrows.append(obj)

    padrows.sort(key = lambda x: x["rid"])

    stacks = {}
    for stack in range(5):
        stacks[stack] = {
            'minZ': math.inf,
            'maxZ': -math.inf,
            'minR': math.inf,
            'maxR': -math.inf
        }

    rangeX = 0

    modules = []
    for module in data:
        obj = {}
        obj["rid"] = (module["stack"] * 6 + module["layer"]) * 16
        obj["stk"] = module["stack"]

        obj["lyr"] = module["layer"]
        obj["row"] = row
        obj["rows"] = module["zegments"]
        obj["p0"] = {
            "x": module["minLocalY"],
            "y": module["minR"],
            "r": module["minR"],
            "z": round(module["minZ"], 2)
        }
        obj["p1"] = {
            "x": module["maxLocalY"],
            "y": module["maxR"],
            "r": module["maxR"],
            "z": round(module["maxZ"], 2)
        }

        modules.append(obj)

        rangeX = max(rangeX, module["maxLocalY"] - module["minLocalY"])

        stack = stacks[module["stack"]]
        stack["minZ"] = round(min(stack["minZ"], module["minZ"]), 2)
        stack["maxZ"] = round(max(stack["maxZ"], module["maxZ"]), 2)
        stack["minR"] = round(min(stack["minR"], module["minR"]), 2)
        stack["maxR"] = round(max(stack["maxR"], module["maxR"]), 2)

    modules.sort(key = lambda x: x["rid"])

    rangeY = 0

    for stack in stacks.values():
        rangeZ = stack["maxZ"] - stack["minZ"]
        rangeR = stack["maxR"] - stack["minR"]

        dimSupermoduleRange = max(rangeZ, rangeR)
        rangeY = min(dimSupermoduleRange, rangeY)
        
        midZ = (stack["maxZ"] + stack["minZ"]) / 2
        stack["bbZ"] = [midZ - dimSupermoduleRange / 2, midZ + dimSupermoduleRange / 2]
        
        midR = (stack["maxR"] + stack["minR"]) / 2
        stack["bbR"] = [midR + dimSupermoduleRange / 2, midR - dimSupermoduleRange / 2]

    
    sectorRange = max(rangeX, rangeY)
    sectorBB = {}
    
    sectorBB["bbX"] = [-sectorRange / 2, +sectorRange / 2]

    midY = (stacks[2]["maxR"] + stacks[2]["minR"]) / 2
    sectorBB["bbY"] = [midY + sectorRange / 2, midY - sectorRange / 2]

    outfile = open("jsroot/components/padrow-dimensions.js", "w")
    print("function getPadrowDimensions() {\nconst dims = ", file = outfile, end = '')
    json.dump(padrows, outfile, indent = 4)
    print(";\n\n\treturn dims;\n}", file = outfile)
    
    print("\n\n")

    print("function getModuleDimensions() {\nconst dims = ", file = outfile, end = '')
    json.dump(modules, outfile, indent = 4)
    print(";\n\n\treturn dims;\n}", file = outfile)

    print("\n\n")

    print("function getStackDimensions() {\nconst dims = ", file = outfile, end = '')
    json.dump(stacks, outfile, indent = 4)
    print(";\n\n\treturn dims;\n}", file = outfile)

    print("\n\n")

    print("function getSectorDimensions() {\nconst dims = ", file = outfile, end = '')
    json.dump(sectorBB, outfile, indent = 4)
    print(";\n\n\treturn dims;\n}", file = outfile)
        
    outfile.close()

if __name__ == "__main__":
    main("A3:R3", "A4:S33")